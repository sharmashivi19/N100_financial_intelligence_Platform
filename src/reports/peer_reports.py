import os
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

GREEN = PatternFill(
    start_color="90EE90",
    end_color="90EE90",
    fill_type="solid"
)

YELLOW = PatternFill(
    start_color="FFF59D",
    end_color="FFF59D",
    fill_type="solid"
)

RED = PatternFill(
    start_color="FF9999",
    end_color="FF9999",
    fill_type="solid"
)

GOLD = PatternFill(
    start_color="FFD700",
    end_color="FFD700",
    fill_type="solid"
)


def generate_peer_report():

    conn = sqlite3.connect("database/nifty100.db")

    companies = pd.read_sql(
        "SELECT * FROM companies",
        conn
    )

    ratios = pd.read_sql(
        "SELECT * FROM financial_ratios",
        conn
    )

    percentiles = pd.read_sql(
        "SELECT * FROM peer_percentiles",
        conn
    )

    peer_groups = pd.read_excel(
        "data/source_files/peer_groups.xlsx"
    )

    conn.close()
    df = ratios.merge(
        companies,
        left_on="company_id",
        right_on="id",
        how="left"
    )

    df = df.merge(
        peer_groups,
        on="company_id",
        how="left"
    )
    pct = percentiles.pivot_table(
        index=["company_id","year"],
        columns="metric",
        values="percentile_rank"
    ).reset_index()

    pct.columns = [
        c if isinstance(c,str)
        else c[1]
        for c in pct.columns
    ]
    df = df.merge(
        pct,
        on=["company_id","year"],
        how="left",
        suffixes=("","_pct")
    )
    os.makedirs("output",exist_ok=True)

    writer = pd.ExcelWriter(
        "output/peer_comparison.xlsx",
        engine="openpyxl"
    )
    groups = df["peer_group_name"].dropna().unique()

    for group in groups:

        sheet = df[
            df["peer_group_name"]==group
        ].copy()

        sheet.to_excel(
            writer,
            sheet_name=group[:31],
            index=False
        )

    writer.close()
    wb = load_workbook(
        "output/peer_comparison.xlsx"
    )
    for ws in wb.worksheets:

        headers = [
            c.value
            for c in ws[1]
        ]

        for col in range(1,ws.max_column+1):

            name = headers[col-1]

            if name.endswith("_pct"):

                for row in range(2,ws.max_row+1):

                    cell = ws.cell(row,col)

                    if cell.value is None:
                        continue

                    if cell.value >= 0.75:
                        cell.fill = GREEN

                    elif cell.value <= 0.25:
                        cell.fill = RED

                    else:
                        cell.fill = YELLOW
        if "is_benchmark" in headers:

            bench_col = headers.index(
                "is_benchmark"
            )+1

            for row in range(2,ws.max_row+1):

                if ws.cell(
                    row,
                    bench_col
                ).value==1:

                    for c in range(
                        1,
                        ws.max_column+1
                    ):

                        ws.cell(
                            row,
                            c
                        ).fill = GOLD
        last = ws.max_row+1

        ws.cell(last,1).value="Median"
        import statistics

        for col in range(2,ws.max_column+1):

            vals=[]

            for row in range(2,last):

                v=ws.cell(row,col).value

                if isinstance(v,(int,float)):
                    vals.append(v)

            if vals:

                ws.cell(
                    last,
                    col
                ).value=statistics.median(vals)
    wb.save(
        "output/peer_comparison.xlsx"
    )

    print(
        "Peer comparison report generated."
    )