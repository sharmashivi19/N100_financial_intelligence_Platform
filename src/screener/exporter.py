import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


green = PatternFill(
    fill_type="solid",
    fgColor="90EE90"
)

red = PatternFill(
    fill_type="solid",
    fgColor="FFC7CE"
)


def export_screener(results, output_file):

    # Write Excel
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        for sheet_name, df in results.items():

            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )

    # Open workbook again
    workbook = load_workbook(output_file)

    for sheet in workbook.sheetnames:

        ws = workbook[sheet]

        headers = {}

        for cell in ws[1]:
            headers[cell.value] = cell.column

        # ROE
        if "return_on_equity_pct" in headers:

            col = headers["return_on_equity_pct"]

            for row in range(2, ws.max_row + 1):

                value = ws.cell(row, col).value

                if value is None:
                    continue

                if value >= 15:
                    ws.cell(row, col).fill = green
                else:
                    ws.cell(row, col).fill = red

        # Debt to Equity
        if "debt_to_equity" in headers:

            col = headers["debt_to_equity"]

            for row in range(2, ws.max_row + 1):

                value = ws.cell(row, col).value

                if value is None:
                    continue

                if value <= 1:
                    ws.cell(row, col).fill = green
                else:
                    ws.cell(row, col).fill = red

        # Free Cash Flow
        if "free_cash_flow_cr" in headers:

            col = headers["free_cash_flow_cr"]

            for row in range(2, ws.max_row + 1):

                value = ws.cell(row, col).value

                if value is None:
                    continue

                if value > 0:
                    ws.cell(row, col).fill = green
                else:
                    ws.cell(row, col).fill = red

        # Revenue CAGR
        if "revenue_cagr_5yr" in headers:

            col = headers["revenue_cagr_5yr"]

            for row in range(2, ws.max_row + 1):

                value = ws.cell(row, col).value

                if value is None:
                    continue

                if value >= 10:
                    ws.cell(row, col).fill = green
                else:
                    ws.cell(row, col).fill = red

    workbook.save(output_file)